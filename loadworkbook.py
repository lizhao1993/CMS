import DataInterface
from openpyxl import load_workbook
from openpyxl.reader import workbook
from openpyxl.writer import worksheet, etree_worksheet
from openpyxl.writer import workbook
from xml.etree.ElementTree import ElementTree

def getStudentsFromWorkbook(filename):
    wb = load_workbook(filename,read_only=True)
    ws = wb.worksheets[0]
    students = []
    exclude = [None,'First Name','Last Name','Email','Units']
    
    for row in ws.rows:
        student = []
        
        for cell in row:
            if cell.value not in exclude:
                student.append(cell.value)
                
        if len(student)!=0:
            students.append(student)
    
    return students


