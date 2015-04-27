#TODOs: getRoster()
#TODO: Get feedback working
#TODO: Export functionality
#TODO: New semester/year functionality

import sys
import os.path
import ast
import loadworkbook
import DataInterface
import xml.etree.ElementTree as ET
from PyQt5.QtCore import *
from PyQt5.QtGui import *
from PyQt5.QtWidgets import *
from CMS1 import Ui_MainWindow
import datetime
from datetime import date

from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.cell import get_column_letter

#--------------------------------------DONE----------------------------------
def populateTableView(model,students):
    """ This is a function for populating the roster;
    populateTableView takes in a nested list of students with their info,
    a model (standardItemModel), and adds data to the TableView; it then
    returns the model.
    This method also adds students to the database so it should only be
    called when the database is empty """    
    row=0;
    # Add students to tableView
    for student in students:
        name = student[0]+" "+student[1]
        email = student[2]
        units = str(student[3])
        db.stuMod(name,"Email",email)
        db.stuMod(name,"Units",units)
        model.setItem(row,0,QStandardItem(name))
        model.setItem(row,1,QStandardItem(email))
        model.setItem(row,2,QStandardItem(units))
        row+=1        
    db.save()
    return model

#--------------------------------------DONE----------------------------------
def populateAttendance(students):
    """ populateattendanceTable takes in a nested list of students with their info
    and adds data to the (default) attendanceTable; it is called when a new
    roster is uploaded """    
    table = ui.attendanceTable
    table.setColumnCount(2)
    table.setHorizontalHeaderItem(0,QTableWidgetItem("Name"))
    table.setHorizontalHeaderItem(1,QTableWidgetItem("Total Unexcused"))
    table.setRowCount(len(students))
    row=0;
    for student in students:
        name = student[0]+" "+student[1]
        db.addStudent(name)
        absences = db.stuCall(name,"Number_of_Absences")
        table.setItem(row,0,QTableWidgetItem(name))
        table.setItem(row,1,QTableWidgetItem(absences))
        row+=1        
    return table

#--------------------------------------DONE----------------------------------
def populateGrades(students):
    """ populateGrades takes in a nested list of students and adds their names
    to the first column in the grades table; it is called when a new roster
    is uploaded """    
    table = ui.gradesTable
    table.setColumnCount(1)
    table.setHorizontalHeaderItem(0,QTableWidgetItem("Name"))
    table.setRowCount(len(students))
    row=0;
    for student in students:
        name1 = QTableWidgetItem(student[0]+" "+student[1])
        table.setItem(row,0,name1)
        row+=1
    return table

#--------------------------------------TODO----------------------------------
def getRoster(self):
    #TODO: Check if the db already has stuff in it and if so, do a
    #      comparison
    """ getRoster is called when the button to import a roster is clicked; it
    reads the imported excel file into a nested list of students and then
    populates the roster, attendanceTable, and the gradesTable """
    
    fname = QFileDialog.getOpenFileName()
    filename = (fname[0])
    students = loadworkbook.getStudentsFromWorkbook(filename)
    # Add students to the table Widget
    table = populateAttendance(students)
    gradesTable = populateGrades(students)
    # TableView needs a model
    model = QStandardItemModel(len(students),3)
    model.setHorizontalHeaderItem(0, QStandardItem("Name"))
    model.setHorizontalHeaderItem(1, QStandardItem("Email"))
    model.setHorizontalHeaderItem(2, QStandardItem("Units"))    
    # Add students to tableView
    model = populateTableView(model,students)    
    # Display the attendanceTable and the tableView
    table.show()
    gradesTable.show()
    ui.rosterView.setModel(model)
    
    db.save()

#--------------------------------------DONE----------------------------------
def cellChangedAttendance(self):
    """
    cellChangedAttendance takes in no inputs and returns no outputs;
    This function is called when a cell in the attendance table is changed;
    it gets the student's name and adds the absence to the database using
    stuAbsence(name)
    """
    col = ui.attendanceTable.currentColumn()
    row = ui.attendanceTable.currentRow()    
    #get student's name
    if col!=0:
        name = ui.attendanceTable.item(row,0)
        if name:
            name = name.text()
            date = ui.attendanceTable.horizontalHeaderItem(col)
            if date:
                date = date.text()
                item = ui.attendanceTable.currentItem()
                if item:
                    item = item.text()
                    db.stuAbsence(name)
                    db.stuMod(name,date,item)                  
                    
                    db.save()

#--------------------------------------DONE----------------------------------
def cellChangedGrades(self):
    """
    cellChangedAttendance takes in no inputs and returns no outputs;
    This function is called when a cell in the grades table is changed;
    it gets the student's name, the name of the homework assingment
    and the value of the changed cell and adds it to the database
    """
    col = ui.gradesTable.currentColumn()
    row = ui.gradesTable.currentRow()  
    if col!=0:
        name = ui.gradesTable.item(row,0) #student name
        if name:
            name = name.text()
            header = ui.gradesTable.horizontalHeaderItem(col)#homework name
            if header:
                header = header.text()
                item = ui.gradesTable.currentItem()#value of changed cell
                if item:
                    item = item.text()
                    db.stuMod(name,header,item)
                    db.save()
                    
#--------------------------------------DONE----------------------------------
def showDialog(self):
    """
    input: none; output: none
    showDialog is called when the user presses the Add Assignment button;
    it gets the name of the assignment, inserts a column in the grades table,
    adds it to the assignments in the database and adds it as a tag in every
    student
    """
    inputDialog = QInputDialog()
    text, ok = inputDialog.getText(ui.add_assignment,"Add Assignment",
                                   "Enter Assignment Name:")
    if ok:
        cols = ui.gradesTable.columnCount()
        ui.gradesTable.insertColumn(cols)        
        ui.gradesTable.setHorizontalHeaderItem(cols,QTableWidgetItem(text))
        
        db.stuAdd(text) #add assignment as tag in each student
        db.addAssignment(text) #add to list of assignments
        db.save()
        
#--------------------------------------DONE----------------------------------
def populateProjTable(model,names):
    """
    input: model for the student in project table view and a list of names
    output: model;
    populateProjTable is called once a new project is added. It creates a two-
    column table that includes the names of the students in the project and the
    number of units each student is registered for. 
    """
    model.setHorizontalHeaderItem(0,QStandardItem("Name"))
    model.setHorizontalHeaderItem(1,QStandardItem("Units"))
    row=0
    for name in names:
        units=db.stuCall(name, "Units")
        model.setItem(row,0,QStandardItem(name))
        model.setItem(row,1,QStandardItem(units))
        row+=1
    return model


#--------------------------------------TODO----------------------------------
def populateFeedTableFromDB(model,groups):
    """
    input: model for the student in project table view and a list of names
    output: model;
    populateProjTable is called once a new project is added. It creates a two-
    column table that includes the names of the students in the project and the
    number of units each student is registered for. 
    """
    model.setHorizontalHeaderItem(0,QStandardItem("Name"))
    model.setHorizontalHeaderItem(1,QStandardItem("Units"))
    row=0
    for name in names:
        units=db.stuCall(name, "Units")
        model.setItem(row,0,QStandardItem(name))
        model.setItem(row,1,QStandardItem(units))
        row+=1
    return model
#--------------------------------------DONE----------------------------------
def projComboBoxFill(dialog):
    """
    input: dialog containing the combo box
    output: the filled comboBox
    projComboBoxFill adds all of the students in the class to the comboBox.
    This function should be called to fill the comboBox in the Add New Project
    dialog. 
    """
    combo = QComboBox(dialog)
    names=db.stuMassCall("Name")
    for name in names:
        combo.addItem(name)
    return combo

#--------------------------------------DONE----------------------------------
def onChanged(self):
    """
    onChanged is called when the value of the SpinBox in the add new project
    dialog is changed by the user
    """
    value = ui.numStudents.value() 
    form=ui.form 
    numAlready=form.rowCount() #number of existing elements
    total=value+4 #the number of elements we want (4 is the base, w/o combo)
    if total>numAlready: #we need boxes
        for i in range(0,total-numAlready):
            combo=projComboBoxFill(ui.dialog)
            form.addRow(combo)
    if total<numAlready: #too many boxes
        for i in range(0,abs(total-numAlready)):
            form.takeAt(6)#removes row 6
    if total==numAlready: #weird thing that we add bc it doesn't add boxes
                          #like we think it should
        combo=projComboBoxFill(ui.dialog)
        form.addRow(combo)
    
#--------------------------------------DONE----------------------------------
def addNewProject(self):
    """
    addNewProject handles the events that need to happen when we click to
    add a new project.
    It creates a dialog for the user to input the name of the project and a
    SpinBox for the user to input the number of students in the projects (see
    onChanged). If the user clicks "OK" in the dialog, accepted is called.
    """
    #TODO: Create model for the student info: name & number of units
    #TODO: Create model for the feedback: date & comment
    #TODO: Add project to the database
    #TODO: Add students to project group
    ui.dialog = QDialog()
    ui.form = QFormLayout(ui.dialog)
    form = ui.form
    #Get the info from the dialog window
    form.addRow(QLabel("Enter project name and students."))
    le = QLineEdit(ui.dialog) #will contain the user input for project names
    form.addRow("Project Name",le)
    #user inputs for the students
    ui.numStudents = QSpinBox(ui.dialog)
    ui.numStudents.setRange(1,10)
    form.addRow("Number of Students",ui.numStudents)
    ui.numStudents.valueChanged.connect(onChanged)
    #Add the OK and Cancel buttons
    buttonBox = QDialogButtonBox(
        QDialogButtonBox.Ok | QDialogButtonBox.Cancel,
        Qt.Horizontal, ui.dialog)
    form.addRow(buttonBox)
    buttonBox.accepted.connect(accepted)
    buttonBox.accepted.connect(ui.dialog.accept)#so it closes when ok pressed
    buttonBox.rejected.connect(ui.dialog.reject)
    
    ui.dialog.exec_()

#--------------------------------------DONE----------------------------------
def addProjectTables(projName):
    numPages = ui.toolBox.count() #the number of existing projects
    combo = ui.chooseProject
    #set up the page to display project data
    ui.page = QWidget()
    ui.page.setGeometry(QRect(0,0,100,30))
    ui.page.setObjectName(projName)
    ui.students = QTableView(ui.page)
    ui.students.setGeometry(QRect(0, 0, 260, 140))
    ui.students.setObjectName("students_"+projName)
    ui.projectFeedback = QTableView(ui.page)
    ui.projectFeedback.setGeometry(QRect(0, 150, 260, 150))
    ui.projectFeedback.setObjectName("projFeedback_"+projName)
    
    #add page to UI & to comboBox
    ui.toolBox.addItem(ui.page,projName)
    combo.addItem(projName)
    return [ui.students,ui.projectFeedback]
    
#--------------------------------------DONE----------------------------------
def accepted():
    """
    accepted handles the events that need to happen once we have the info
    about the project and the students. It gets the Project's name and creates
    a tableView for the students' information and a tableView for the feedback
    on the project. It makes calls to populateProjTable and populateFeedTable.
    It also adds the project group and students in it to the database.
    """
    form=ui.form
    rows=form.count()
    fields=[]#contains names of students

    projNameObj=form.itemAt(2)#returns QLayoutItem
    projName=projNameObj.widget()#should be lineEdit
    projName=projName.text()
    db.addGroup(projName)#add to DB

    for i in range(6,rows):#starts at item 6 in row 4
        layoutObj=form.itemAt(i)
        objWidget=layoutObj.widget()
        comboText=objWidget.currentText()
        fields.append(comboText)
        db.groStuAdd(projName,comboText)#add student to group

    db.save()
    
    tables=addProjectTables(projName)
    students=tables[0]
    projectFeedback=tables[1]
    
    #create model for student & units:
    model=QStandardItemModel(len(fields),2)
    model=populateProjTable(model,fields)
    students.show()
    students.setModel(model)
    
    #create model for project feedback & dates
    feedmodel=QStandardItemModel(1,2)
    feedmodel.setHorizontalHeaderItem(0,QStandardItem("Date"))
    feedmodel.setHorizontalHeaderItem(1,QStandardItem("Feedback"))
    projectFeedback.show()
    projectFeedback.setModel(feedmodel)

#--------------------------------------DONE?----------------------------------
def submitFeedback(self):
    proj=ui.chooseProject.currentText()#get project name from ComboBox
    ddate=ui.projectDateEdit.date()#get date from DateEdit
    ddate=ddate.toString()
    points=ui.weeklyPoints.currentText()#get weekely points from ComboBox
    text=ui.feedBackText.toPlainText()#get text from TextEdit

    project=ui.page.findChild(QTableView,"projFeedback_"+proj)#should return a tableView
    feedModel=project.model()
    toAdd=[QStandardItem(ddate),QStandardItem(text)]
    
    if feedModel==None:
        feedModel=QStandardItemModel()
    feedModel.appendRow(toAdd)
    project.setModel(feedModel)
    project.show()
    
    db.groAdd(proj,ddate,points)#add weekly points and date of feedback
    db.groCommentMode(proj,ddate,text)
    db.save()    

#--------------------------------------DONE----------------------------------
def addTodaysDate(self):
    """
    addTodaysDate is called when the user clicks the Add Date button on the
    attendance tab. It creates an InputDialog that get's the value of the date
    the user wishes to add to the attendance table; it adds the date to the
    database under Dates and to each student with the value of the date to "Y" 
    (default).
    """
    inputDialog = QInputDialog()
    today, ok = inputDialog.getText(ui.add_assignment,"Add Date",
                                   "Enter Date:")
    if ok:
        colnum = ui.attendanceTable.columnCount()
        ui.attendanceTable.insertColumn(colnum)
        ui.attendanceTable.setHorizontalHeaderItem(colnum,QTableWidgetItem(today))
        db.stuAdd(today,"Y")
        db.addDate(today)        
        db.save()
        
        rows = ui.attendanceTable.rowCount()
        for i in range(0,rows):
            ui.attendanceTable.setItem(i,colnum,QTableWidgetItem("Y"))

        
#--------------------------------------DONE----------------------------------
def populateRosterFromDB(model,names):
    """
    input: model for the roster tableView and the list of student names.
    populateRosterFromDB gets the values of the students' emails and units
    and adds them to the model for each student.
    """
    emails = db.stuMassCall("Email")
    units = db.stuMassCall("Units")
    for row in range(0,len(names)):
        model.setItem(row,0,QStandardItem(names[row]))
        model.setItem(row,1,QStandardItem(emails[row]))
        model.setItem(row,2,QStandardItem(units[row]))
    
#--------------------------------------DONE----------------------------------
def populateAttendanceFromDB(names):
    """
    input: list of student names.
    populateAttendanceFromDB finds all the dates for the attendance to display
    them in the attendance table. It also displays the total unexcused absences
    for each student.
    """
    dates = db.findDates()    
    table = ui.attendanceTable
    table.setColumnCount(len(dates)+2)
    table.setRowCount(len(names))
    table.setHorizontalHeaderItem(0,QTableWidgetItem("Name"))
    table.setHorizontalHeaderItem(1,QTableWidgetItem("Total Unexcused"))
    col = table.columnCount()-1
    rows=len(names)

    if len(dates)!=0:
        for date in dates:
            table.setHorizontalHeaderItem(col,QTableWidgetItem(date))
            for i in range(0,rows):
                table.setItem(i,0,QTableWidgetItem(names[i]))
                absences = db.stuCall(names[i],"Number_of_Absences")
                table.setItem(i,1,QTableWidgetItem(absences))
                #look up student's attendance on that date
                att=db.stuCall(names[i],date)
                table.setItem(i,col,QTableWidgetItem(att))
            col=col-1

    else:
         for i in range(0,rows):
            table.setItem(i,0,QTableWidgetItem(names[i]))
            absences = db.stuCall(names[i],"Number_of_Absences")
            table.setItem(i,1,QTableWidgetItem(absences))
                
    return table

#--------------------------------------DONE----------------------------------
def populateGradesFromDB(names):
    """
    input: list of student names
    populateGradesFromDB takes in a nested list of students and adds their
    names to the first column in the grades table. It also adds their
    grade for each assingment.
    """    
    table = ui.gradesTable
    assignments = db.findHW()
    table.setColumnCount(len(assignments)+1)
    table.setHorizontalHeaderItem(0,QTableWidgetItem("Name"))
    table.setRowCount(len(names))
    row=0
    for student in names:
        name = names[row]
        table.setItem(row,0,QTableWidgetItem(name))
        col=1
        for hw in assignments:
            table.setHorizontalHeaderItem(col,QTableWidgetItem(hw))
            grade = db.stuCall(name,hw) #get student's grade
            table.setItem(row,col,QTableWidgetItem(grade))
            col+=1
        row+=1
    return table

#--------------------------------------TODO----------------------------------
def export():
    """ export saves the student name and final grades into a excel file"""
    names = db.stuMassCall("Name")
    #call the grade function for each student 
    for name in names:
        db.stuGrade(name)
    #get the list of grade from database
    finalgrades = db.stuMassCall("Grade")
     
    #make an excel workbook

    wb = Workbook()
    dest_filename = 'Export.xlsx'

    #make a worksheet
    ws1 = wb.active
    ws1.title = "Grade"

    #put in the names and grades
    #First value is header
    ws1.cell(row=1, column=1).value = "Name"
    r = 2
    for name in names:
        ws1.cell(row=r, column=1).value = name
        r += 1

    ws1.cell(row=1, column=2).value = "Grade"
    j = 2
    for grade in finalgrades:
        ws1.cell(row=j, column=2).value = grade
        j += 1

    #save the file
    wb.save(filename = dest_filename)
    
            
            
        
#--------------------------------------TODO----------------------------------
if __name__=="__main__":
    app = QApplication(sys.argv)
    window = QMainWindow()
    ui = Ui_MainWindow()
    ui.setupUi(window)

    filename = "database.xml"
    
    if os.path.isfile(filename):
        db = DataInterface.DataInterface(filename)
        
        names = db.stuMassCall("Name")
        #Roster, Attendance, and Grades all get filled here:
        model = QStandardItemModel(len(names),3)
        model.setHorizontalHeaderItem(0, QStandardItem("Name"))
        model.setHorizontalHeaderItem(1, QStandardItem("Email"))
        model.setHorizontalHeaderItem(2, QStandardItem("Units"))
        
        populateRosterFromDB(model,names)
        ui.rosterView.setModel(model)
        
        attendanceTable = populateAttendanceFromDB(names)
        gradesTable = populateGradesFromDB(names)
        
        #Project stuff gets filled here:
        groups=db.findAllGroups()#contains the project names
        for group in groups:
            tables=addProjectTables(group)
            students=tables[0]
            projectFeedback=tables[1]
            
            studentsInGroup=db.findGroup(group).find("Students").attrib["info"]
            studentsInGroup=ast.literal_eval(studentsInGroup)#this converts
                            #the string representation of the list into an
                            #actual list
            studentModel=QStandardItemModel(len(studentsInGroup),2)
            studentModel=populateProjTable(studentModel,studentsInGroup)
            students.setModel(studentModel)
            students.show()
        ui.weeklyPoints.addItem("1")
        ui.weeklyPoints.addItem("2")
        ui.weeklyPoints.addItem("3")

            
    else:
        db = DataInterface.DataInterface()
	

    ui.pushButton.clicked.connect(getRoster)
    ui.add_assignment.clicked.connect(showDialog)
    ui.addProjectBttn.clicked.connect(addNewProject)
    ui.addDateButton.clicked.connect(addTodaysDate)
    ui.attendanceTable.cellChanged.connect(cellChangedAttendance)
    ui.gradesTable.cellChanged.connect(cellChangedGrades)    
    ui.export_2.clicked.connect(export)
    ui.submitFeedback.clicked.connect(submitFeedback)

    ui


    window.show()
    sys.exit(app.exec_())

